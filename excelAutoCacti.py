from pickle import FALSE
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import numbers
from percentileFinderCacti import findValue
from daysNumber import daysNumber


# ensure the filename, sequence, and cid.
def createReport(selection, startPicture, endPicture):
    xlsFileName = input('copy your excel file name: ')
    endOfMonth = True
    fileName = xlsFileName + '.xlsx'
    wb = load_workbook(fileName)
    if selection == 1 :
        cid = 8254
    if selection == 2 :
        cid = 10022
    if selection == 3 : #HE#2
        cid = 11301
    if selection == 4 :
        cid = 11356
    if selection == 5 :
        cid = 10690
    if selection == 6 :
        cid = 11213
    if selection == 7: 
        cid = 11211
    if selection == 8: 
        cid = 11357
    if selection == 9: 
        cid = 7424
    if selection == 10: 
        cid = 8381 
    if selection == 11:
        cid = 11869
    if selection == 12:
        cid = 11872
    if selection == 13:
        cid = 10046
    if selection == 14:
        cid = 9502
    if selection == 15:
        cid = 11953


    print('This will be working on row', startPicture, 'until', endPicture)
    cont = input(
        'Please ensure your CID is {} , and please check the filename. Are those okay? y/n '.format(cid))
    if cont == 'n':
        print('Please change on the your cid and/or file.')
        exit()

    ws = wb.active


    p2e = pixels_to_EMU
    c2e = cm_to_EMU

    size = XDRPositiveSize2D(p2e(295.68), p2e(131.52))


    # Calculated number of cells width or height from cm into EMUs
    def cellh(x): return c2e((x * 49.77)/99)
    def cellw(x): return c2e((x * (18.65-1.71))/10)

    # Want to place image in row 5 (6 in excel), column 2 (C in excel)
    # Also offset by half a column.
    column = 10
    coloffset = cellw(0.15)
    row = 36
    rowoffset = cellh(0.7)

    for i in range(startPicture, endPicture):
        # Define the name by cid:
        if cid == 11356: #0
            newNameFile = 'PGAS.VAL.43.01 - {}.png'.format(i)
        if cid == 8254: #1
            newNameFile = 'PGAS.VAL.33.01#1 - {}.png'.format(i)
        if cid == 10022: #2
            newNameFile = 'PGAS.VAL.33.01#2 - {}.png'.format(i)
        if cid == 11301: #3
            newNameFile = 'PGAS.VAL.33.04 - {}.png'.format(i)
        if cid == 11356: #4
            newNameFile = 'PGAS.VAL.43.01 - {}.png'.format(i)
        if cid == 10690: #5
            newNameFile = 'PGAS.VAL.43.02 - {}.png'.format(i)
        if cid == 11213: #6
            newNameFile = 'PGAS.VAL.43.03 - {}.png'.format(i)
        if cid == 11211: #7
            newNameFile = 'PGAS.VAL.43.04 - {}.png'.format(i)
        if cid == 11357: #8
            newNameFile = 'PGAS.VAL.43.05 - {}.png'.format(i)
        if cid == 7424: #9
            newNameFile = 'NTT - {}.png'.format(i)
        if cid == 8381: #10
            newNameFile = 'TELIA - {}.png'.format(i)
        if cid == 11869: #11
            newNameFile = 'PGAS.VAL.33.05 - {}.png'.format(i) # HE#3
        if cid == 11872: #12
            newNameFile = 'PGAS.VAL.33.06 - {}.png'.format(i) # HE#4
        if cid == 10046: #13
            newNameFile = 'PCCW - {}.png'.format(i)
        if cid == 9502: #14
            newNameFile = 'TATA - {}.png'.format(i)
        if cid == 11953: #15
            newNameFile = 'PGAS.VAL.43.06 - {}.png'.format(i)

        img = Image('./graph/{}'.format(newNameFile))
        # This is for table

        try:
            if newNameFile:
                result = findValue('{}'.format(newNameFile))
                result1 = result[0][0] # Avg. Inbound
                result2 = result[0][1].replace('.',',') # Avg. Outbound
                result3 = result[1][0].replace('.',',') # Max. Inbound
                result4 = result[1][1].replace('.',',') # Max. Outbound
                result5 = result[2][0].replace('.',',') # Percentile
                cellRow = 36 + i
                print(result1, result2, result3, result4, result5)
                ws['D{}'.format(cellRow)] = result1
                ws['E{}'.format(cellRow)]  = result2  
                ws['F{}'.format(cellRow)] = result3
                ws['G{}'.format(cellRow)]  = result4  
                ws['H{}'.format(cellRow)]  = result5  
                
                ws['D{}'.format(cellRow)].number_format = '#,##'
                ws['E{}'.format(cellRow)].number_format = '#,##'
                ws['F{}'.format(cellRow)].number_format = '#,##'
                ws['G{}'.format(cellRow)].number_format = '#,##'
                ws['H{}'.format(cellRow)].number_format = '#,##'
            # insertValue = findValue('{}.1'.format(i))
            # ws['D{}'.format(row + i)] = insertValue[0]
            # ws['F{}'.format(row + i)] = insertValue[1]

        except:
            print('data unavailable')

        h, w = img.height, img.width

        if i != 1:
            row = row + 1

        marker1 = AnchorMarker(col=column, colOff=coloffset,
                            row=row, rowOff=rowoffset)

        img.anchor = OneCellAnchor(_from=marker1, ext=size)
        ws.add_image(img)
        print('working for {} table'.format(i))

    if endOfMonth == True:

        try:
            # Test
            # newNameFile = 'NTT - {}.png'
            if newNameFile:
                newNameFile = newNameFile[:-7]
                result = findValue('{} 32.png'.format(newNameFile))
                # Average field
                result1 = result[0][0] # Avg. Inbound
                result2 = result[0][1].replace('.',',') # Avg. Outbound
                result3 = result[1][0].replace('.',',') # Max. Inbound
                result4 = result[1][1].replace('.',',') # Max. Outbound
                result5 = result[2][0].replace('.',',') # Percentile
                cellRow = 69
                print(result1, result2, result3, result4, result5)
                ws['D{}'.format(cellRow)] = result1
                ws['E{}'.format(cellRow)]  = result2  
                ws['F{}'.format(cellRow)] = result3
                ws['G{}'.format(cellRow)]  = result4  
                ws['H{}'.format(cellRow)]  = result5  
                

        except:
            print('data not found')

        img = Image('./graph/{} 32.png'.format(newNameFile))
        # This is for table
        marker1 = AnchorMarker(col=10, colOff=coloffset, row=68, rowOff=rowoffset)

        img.anchor = OneCellAnchor(_from=marker1, ext=size)
        ws.add_image(img)
        print('32 is added')
    wb.save('[completed]{}'.format(fileName))

createReport(14, 1, daysNumber(2022, 10) + 1)

# done all
